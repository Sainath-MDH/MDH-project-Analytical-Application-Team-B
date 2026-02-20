from openpyxl import load_workbook

try:
    wb = load_workbook(filename=r'c:\Users\saina\OneDrive\Documents\Bike Theft\MDH-project-Analytical-Application-Team-B\berlin_LOR_hierarchy.xlsx', read_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}")
    for row in ws.iter_rows(max_row=5, values_only=True):
        print(row)
    
    # Try to find districts
    # Usually districts are in a column. Let's look for known Berlin districts.
    known_districts = ["Mitte", "Pankow", "Lichtenberg", "Spandau"]
    district_idx = None
    header = next(ws.iter_rows(max_row=1, values_only=True))
    print(f"Header: {header}")
    
    # Check rows for districts
    districts_found = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        for val in row:
            if val in [
                "Charlottenburg-Wilmersdorf", "Friedrichshain-Kreuzberg", "Lichtenberg", 
                "Marzahn-Hellersdorf", "Mitte", "Neukölln", "Pankow", 
                "Reinickendorf", "Spandau", "Steglitz-Zehlendorf", 
                "Tempelhof-Schöneberg", "Treptow-Köpenick"
            ]:
                districts_found.add(val)
    
    print("\nDistricts found in Excel:")
    for d in sorted(list(districts_found)):
        print(f"- {d}")
        
except Exception as e:
    print(f"Error: {e}")
